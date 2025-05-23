import openpyxl
import os

# 获取当前脚本所在目录
base_dir = os.path.dirname(os.path.abspath(__file__))
input_filename = '1.xlsx'
output_filename = 'xxx_filtered_debug.xlsx' # New output file to avoid confusion
input_path = os.path.join(base_dir, input_filename)
output_path = os.path.join(base_dir, output_filename)

print(f"输入文件: {input_path}")
print(f"输出文件: {output_path}")

# 检查输入文件是否存在
if not os.path.exists(input_path):
    print(f"错误: 输入文件 {input_path} 未找到。请确保文件存在于正确的位置。")
    exit()

# 加载Excel文件
try:
    wb = openpyxl.load_workbook(input_path)
    ws = wb.active
except Exception as e:
    print(f"错误: 加载Excel文件 {input_path} 失败: {e}")
    exit()

# 输出表头
header = [cell.value for cell in ws[1]]
print('表头:', header)

# 找到“单位级别”所在的列索引
level_column_name = '单位级别'
try:
    level_col_idx_0based = header.index(level_column_name) # 0-based index
except ValueError:
    print(f"错误：未在表头 {header} 中找到列名 '{level_column_name}'。请检查表头是否正确。")
    exit()

print(f"'{level_column_name}' 列位于索引 {level_col_idx_0based} (0-based)。")

# 新建工作簿和sheet
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# 写入表头到新sheet
for col_idx, header_value in enumerate(header, start=1):
    new_ws.cell(row=1, column=col_idx, value=header_value)

# 筛选单位级别为1级的行
new_sheet_row_idx = 2
data_found_and_written = False

print("\n开始逐行处理数据 (从第二行开始):")
for r_idx, row_cells in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
    if len(row_cells) <= level_col_idx_0based:
        print(f"警告: 第 {r_idx} 行的列数不足 ({len(row_cells)} 列)，无法获取 '{level_column_name}'。跳过此行。")
        continue

    level_cell = row_cells[level_col_idx_0based]
    level_value_original = level_cell.value

    print(f"  行 {r_idx}: 原始 '{level_column_name}' 值: '{level_value_original}' (类型: {type(level_value_original)})")

    if level_value_original is not None:
        level_value_str = str(level_value_original).strip()
        processed_val = level_value_str.replace('１', '1').replace('级', '').replace(' ', '')
        print(f"    处理后的值: '{processed_val}'")

        if processed_val == '1':
            print(f"    匹配成功! 将此行写入输出文件。")
            for c_idx, cell_in_row in enumerate(row_cells, start=1):
                new_ws.cell(row=new_sheet_row_idx, column=c_idx, value=cell_in_row.value)
            new_sheet_row_idx += 1
            data_found_and_written = True
        else:
            print(f"    不匹配 '1'。")
    else:
        print(f"    原始值为 None，跳过。")

# 保存新文件
try:
    new_wb.save(output_path)
    if data_found_and_written:
        print(f"\n已生成 {output_path}，包含筛选后的数据。")
    else:
        print(f"\n处理完成，但在 {input_filename} 中未找到符合条件 (单位级别为 '1级') 的数据。")
        print(f"输出文件 {output_path} 已生成，可能只包含表头。")
except Exception as e:
    print(f"错误: 保存输出文件 {output_path} 失败: {e}")

print("\n脚本执行完毕。请检查上面的打印输出以了解详细的筛选过程和结果。")