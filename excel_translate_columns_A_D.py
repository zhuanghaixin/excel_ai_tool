import openpyxl
from deep_translator import GoogleTranslator
import os
from tqdm import tqdm
import string

def column_letter_to_index(column_letter):
    """将列字母转换为索引（A=0, B=1, ...）"""
    return string.ascii_uppercase.index(column_letter.upper())

# 获取当前脚本所在目录
base_dir = os.path.dirname(os.path.abspath(__file__))

# 用户输入文件名
input_filename = input("请输入要翻译的Excel文件名（如test.xlsx）: ")
if not input_filename.endswith('.xlsx'):
    input_filename += '.xlsx'
input_path = os.path.join(base_dir, input_filename)

# 检查文件是否存在
if not os.path.exists(input_path):
    print(f"错误：找不到文件 {input_filename}。请确认文件名并重试。")
    exit()

# 用户指定要翻译的列
print("\n请指定要翻译的两列：")
zh_to_en_column = input("哪一列需要从中文翻译成英文（输入列名，如A、B等）: ").strip().upper()
en_to_zh_column = input("哪一列需要从英文翻译成中文（输入列名，如A、B等）: ").strip().upper()

# 验证列名有效性
valid_columns = list(string.ascii_uppercase)
if zh_to_en_column not in valid_columns or en_to_zh_column not in valid_columns:
    print("错误：请输入有效的列名（A-Z）")
    exit()

# 将列名转换为索引
zh_to_en_idx = column_letter_to_index(zh_to_en_column)
en_to_zh_idx = column_letter_to_index(en_to_zh_column)

# 输出确认信息
print(f"\n将翻译文件 {input_filename} 的：")
print(f"- {zh_to_en_column}列（从中文翻译成英文）")
print(f"- {en_to_zh_column}列（从英文翻译成中文）")
confirm = input("请确认(y/n): ").strip().lower()
if confirm != 'y':
    print("已取消操作")
    exit()

# 初始化两个翻译器
zh_to_en_translator = GoogleTranslator(source='zh-CN', target='en')
en_to_zh_translator = GoogleTranslator(source='en', target='zh-CN')

# 自动生成输出文件名
name, ext = os.path.splitext(input_filename)
output_filename = f'{name}_translated{ext}'
output_path = os.path.join(base_dir, output_filename)

# 加载Excel文件
print(f"\n正在加载 {input_filename}...")
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# 获取数据的最大行列
max_row = ws.max_row
max_col = ws.max_column

# 创建新工作簿
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

print(f"\n开始翻译{zh_to_en_column}列(中→英)和{en_to_zh_column}列(英→中)...")

# 处理每一行
for row_idx in tqdm(range(1, max_row + 1), desc='翻译进度'):
    new_row = []
    
    # 处理每一列
    for col_idx in range(1, max_col + 1):
        col_0_based = col_idx - 1  # 转为0-based索引
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        new_row.append(cell_value)
        
        # 处理中文→英文翻译
        if col_0_based == zh_to_en_idx:
            if cell_value is not None and isinstance(cell_value, str):
                try:
                    translated = zh_to_en_translator.translate(cell_value)
                    if row_idx == 1:  # 如果是表头
                        translated = f"{cell_value}_en"
                except Exception as e:
                    translated = ''
                    print(f"\n翻译失败: {cell_value}, 错误: {str(e)}")
            else:
                translated = ''
            new_row.append(translated)
            
        # 处理英文→中文翻译
        elif col_0_based == en_to_zh_idx:
            if cell_value is not None and isinstance(cell_value, str):
                try:
                    translated = en_to_zh_translator.translate(cell_value)
                    if row_idx == 1:  # 如果是表头
                        translated = f"{cell_value}_zh"
                except Exception as e:
                    translated = ''
                    print(f"\n翻译失败: {cell_value}, 错误: {str(e)}")
            else:
                translated = ''
            new_row.append(translated)
    
    # 将行写入新表格
    new_ws.append(new_row)

# 保存新文件
new_wb.save(output_path)
print(f'\n翻译完成，已生成 {output_path}')
print(f'- {zh_to_en_column}列：中文→英文')
print(f'- {en_to_zh_column}列：英文→中文') 