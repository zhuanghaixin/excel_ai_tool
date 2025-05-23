import openpyxl
from deep_translator import GoogleTranslator, MyMemoryTranslator, BaiduTranslator
import os
from tqdm import tqdm
import string
import time
from collections import defaultdict
import sys
import argparse
import re
import pandas as pd
import csv
import concurrent.futures
import datetime

def column_letter_to_index(column_letter):
    """将列字母转换为索引（A=0, B=1, ...）"""
    return string.ascii_uppercase.index(column_letter.upper())

def index_to_column_letter(index):
    """将索引转换为列字母（0=A, 1=B, ...）"""
    return string.ascii_uppercase[index]

# 批量翻译函数
def batch_translate(translator, texts, batch_size=10, delay=1):
    """批量翻译文本，减少API调用次数"""
    if not texts:
        return []
        
    # 去重以减少翻译量
    unique_texts = list(set(texts))
    print(f"需要翻译 {len(texts)} 个单元格，去重后 {len(unique_texts)} 个唯一文本")
    
    # 创建翻译缓存
    translation_cache = {}
    results = []
    
    # 将文本分批处理
    batches = [unique_texts[i:i+batch_size] for i in range(0, len(unique_texts), batch_size)]
    
    for i, batch in enumerate(tqdm(batches, desc="批次进度")):
        try:
            # 将多个文本合并为一个长文本，用特殊分隔符隔开
            combined_text = " ||| ".join(batch)
            
            # 翻译合并后的文本
            translated = translator.translate(combined_text)
            
            # 根据同样的分隔符拆分翻译结果
            translated_parts = translated.split(" ||| ")
            
            # 如果翻译结果数量与原文本不符，则逐个翻译
            if len(translated_parts) != len(batch):
                print(f"\n批量翻译结果异常，切换为逐个翻译...")
                for text in batch:
                    try:
                        trans = translator.translate(text)
                        translation_cache[text] = trans
                    except Exception as e:
                        print(f"翻译失败: {text}, 错误: {str(e)}")
                        translation_cache[text] = text  # 失败时用原文
            else:
                # 缓存翻译结果
                for j, text in enumerate(batch):
                    translation_cache[text] = translated_parts[j]
            
            # 批次之间添加延迟，避免API限制
            if i < len(batches) - 1 and delay > 0:
                time.sleep(delay)
                
        except Exception as e:
            print(f"\n批次翻译失败，切换为逐个翻译...")
            for text in batch:
                try:
                    trans = translator.translate(text)
                    translation_cache[text] = trans
                except Exception as e:
                    print(f"翻译失败: {text}, 错误: {str(e)}")
                    translation_cache[text] = text  # 失败时用原文
    
    # 根据原始顺序返回翻译结果
    for text in texts:
        results.append(translation_cache.get(text, text))
    
    return results

def display_header():
    """显示应用程序标题"""
    print("="*60)
    print("Excel自动翻译工具 - 支持中英文互译")
    print("="*60)

def get_excel_file():
    """交互式选择Excel文件"""
    # 获取当前目录下所有Excel文件
    excel_files = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.xls'))]
    
    if not excel_files:
        print("当前目录下没有找到Excel文件(.xlsx或.xls)")
        file_path = input("请输入Excel文件的完整路径: ")
        if not file_path.endswith(('.xlsx', '.xls')):
            file_path += '.xlsx'
    else:
        # 显示可选的Excel文件
        print("\n当前目录下的Excel文件:")
        for i, file in enumerate(excel_files, 1):
            print(f"{i}. {file}")
        
        # 让用户选择文件
        choice = input("\n请选择要翻译的Excel文件序号（或输入其他文件路径）: ")
        try:
            index = int(choice) - 1
            if 0 <= index < len(excel_files):
                file_path = excel_files[index]
            else:
                print("无效的选择，请输入有效的序号或文件路径")
                return None
        except ValueError:
            # 用户可能直接输入了文件路径
            file_path = choice
            if not file_path.endswith(('.xlsx', '.xls')):
                file_path += '.xlsx'
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"错误：找不到文件 {file_path}。请确认文件名并重试。")
        return None
        
    return file_path

def select_translator():
    """交互式选择翻译API"""
    print("\n请选择翻译API（某些地区无法使用Google翻译）:")
    print("1. MyMemory翻译 (免费，无需密钥，推荐首选)")
    print("2. Google翻译 (部分地区可能无法访问)")
    print("3. 百度翻译 (需要API ID和密钥)")

    translators = {}
    
    while True:
        api_choice = input("请选择翻译API (1-3): ").strip()
        
        try:
            if api_choice == '1':  # MyMemory翻译
                # MyMemory对于中文使用"zh-CN"，对于英文使用"en-GB"
                translators['zh_to_en'] = MyMemoryTranslator(source='zh-CN', target='en-GB')
                translators['en_to_zh'] = MyMemoryTranslator(source='en-GB', target='zh-CN')
                break
                
            elif api_choice == '2':  # Google翻译
                translators['zh_to_en'] = GoogleTranslator(source='zh-CN', target='en')
                translators['en_to_zh'] = GoogleTranslator(source='en', target='zh-CN')
                break
                
            elif api_choice == '3':  # 百度翻译
                baidu_appid = input("请输入百度翻译API的APP ID: ")
                baidu_key = input("请输入百度翻译API的密钥: ")
                
                try:
                    translators['zh_to_en'] = BaiduTranslator(
                        appid=baidu_appid,
                        appkey=baidu_key,
                        source='zh', 
                        target='en'
                    )
                    translators['en_to_zh'] = BaiduTranslator(
                        appid=baidu_appid,
                        appkey=baidu_key,
                        source='en',
                        target='zh'
                    )
                    break
                except Exception as e:
                    print(f"初始化百度翻译API失败: {e}")
                    continue
            else:
                print("无效选择，请重新输入")
                
        except Exception as e:
            print(f"初始化翻译API失败: {e}")
            print("请重新选择翻译API")
    
    return translators

def parse_column_input(column_input):
    """解析用户输入的列名，支持多种分隔符和格式"""
    if not column_input:
        return []
    
    # 移除所有空格
    column_input = column_input.replace(' ', '')
    
    # 使用正则表达式匹配列名（支持A、B、C或A,B,C或A，B，C等格式）
    columns = re.findall(r'[A-Za-z]', column_input)
    return [col.upper() for col in columns]

def select_columns(wb):
    """交互式选择要翻译的列，支持多列选择"""
    ws = wb.active
    max_col = ws.max_column
    
    # 显示表头信息
    print("\n文件中的列信息:")
    headers = {}
    for col_idx in range(1, max_col + 1):
        col_letter = index_to_column_letter(col_idx - 1)
        header_value = ws.cell(row=1, column=col_idx).value
        headers[col_letter] = header_value
        print(f"{col_letter}. {header_value}")
    
    print("\n请指定要翻译的列：(如不需要某方向翻译，请直接按回车跳过)")
    print("支持多列输入，例如: A,B,C 或 A、B、C 或 ABC")
    
    zh_to_en_input = input("哪些列需要从【中文翻译成英文】（输入列字母，如A、B等，不需要则直接按回车）: ").strip()
    en_to_zh_input = input("哪些列需要从【英文翻译成中文】（输入列字母，如A、B等，不需要则直接按回车）: ").strip()
    
    # 解析用户输入的列名
    zh_to_en_columns = parse_column_input(zh_to_en_input)
    en_to_zh_columns = parse_column_input(en_to_zh_input)
    
    # 验证至少有一个翻译方向
    if not zh_to_en_columns and not en_to_zh_columns:
        print("错误：请至少指定一个翻译方向")
        return None, None
    
    # 验证列名有效性
    valid_columns = list(string.ascii_uppercase[:26])  # 只支持A-Z
    zh_to_en_indices = []
    en_to_zh_indices = []
    
    # 验证中文→英文的列
    for col in zh_to_en_columns:
        if col not in valid_columns:
            print(f"错误：'{col}'不是有效的列名（A-Z）")
            return None, None
        col_idx = column_letter_to_index(col)
        if col_idx >= max_col:
            print(f"错误：列{col}超出文件范围，文件只有{max_col}列")
            return None, None
        zh_to_en_indices.append(col_idx)
    
    # 验证英文→中文的列
    for col in en_to_zh_columns:
        if col not in valid_columns:
            print(f"错误：'{col}'不是有效的列名（A-Z）")
            return None, None
        col_idx = column_letter_to_index(col)
        if col_idx >= max_col:
            print(f"错误：列{col}超出文件范围，文件只有{max_col}列")
            return None, None
        en_to_zh_indices.append(col_idx)
    
    # 检查是否有重复选择的列
    duplicate_cols = set(zh_to_en_columns) & set(en_to_zh_columns)
    if duplicate_cols:
        print(f"警告：列 {', '.join(duplicate_cols)} 同时被选择为中文→英文和英文→中文翻译")
        confirm_duplicate = input("是否继续？(y/n): ").strip().lower()
        if confirm_duplicate != 'y':
            return None, None
    
    # 输出确认信息
    print(f"\n将翻译：")
    if zh_to_en_columns:
        print("从中文翻译成英文的列:")
        for col in zh_to_en_columns:
            print(f"- {col}列 ({headers.get(col, '未知表头')})")
    
    if en_to_zh_columns:
        print("从英文翻译成中文的列:")
        for col in en_to_zh_columns:
            print(f"- {col}列 ({headers.get(col, '未知表头')})")
    
    confirm = input("\n请确认(y/n): ").strip().lower()
    if confirm != 'y':
        print("已取消操作")
        return None, None
    
    return zh_to_en_indices, en_to_zh_indices

def get_batch_size():
    """交互式设置批量翻译大小"""
    default_batch_size = 10
    try:
        custom_batch = input(f"\n设置批量翻译大小（默认每批{default_batch_size}个，较大的值翻译更快但可能失败）: ").strip()
        if custom_batch:
            batch_size = int(custom_batch)
        else:
            batch_size = default_batch_size
    except:
        print(f"使用默认批量大小: {default_batch_size}")
        batch_size = default_batch_size
    
    return batch_size

def translate_excel_file(input_path, translators, zh_to_en_indices, en_to_zh_indices, batch_size=10):
    """执行Excel文件翻译，支持多列翻译"""
    # 记录开始时间
    start_time = time.time()
    
    # 自动生成输出文件名
    name, ext = os.path.splitext(os.path.basename(input_path))
    output_filename = f'{name}_translated{ext}'
    output_path = os.path.join(os.path.dirname(input_path), output_filename)
    
    # 加载Excel文件
    print(f"\n正在加载 {os.path.basename(input_path)}...")
    
    # 使用pandas读取Excel文件，而不是openpyxl，可以更方便地处理列的插入
    df = pd.read_excel(input_path)
    max_row, max_col = df.shape
    
    print(f"文件加载完成，共有 {max_row} 行，{max_col} 列")
    
    # 收集所有需要翻译的文本
    zh_to_en_columns = []
    en_to_zh_columns = []
    zh_to_en_texts = []
    en_to_zh_texts = []
    
    # 将索引转换为列名
    for idx in zh_to_en_indices:
        zh_to_en_columns.append(df.columns[idx])
    
    for idx in en_to_zh_indices:
        en_to_zh_columns.append(df.columns[idx])
    
    # 收集要翻译的文本
    print("正在收集需要翻译的文本...")
    for col in zh_to_en_columns:
        # 跳过表头，只翻译内容
        texts = df[col].iloc[1:].dropna().astype(str).tolist()
        zh_to_en_texts.extend(texts)
    
    for col in en_to_zh_columns:
        # 跳过表头，只翻译内容
        texts = df[col].iloc[1:].dropna().astype(str).tolist()
        en_to_zh_texts.extend(texts)
    
    # 打印开始翻译的信息
    translate_info = []
    if zh_to_en_indices:
        zh_to_en_cols = [index_to_column_letter(idx) for idx in zh_to_en_indices]
        translate_info.append(f"{','.join(zh_to_en_cols)}列(中→英)")
    if en_to_zh_indices:
        en_to_zh_cols = [index_to_column_letter(idx) for idx in en_to_zh_indices]
        translate_info.append(f"{','.join(en_to_zh_cols)}列(英→中)")
    print(f"\n开始翻译{' 和 '.join(translate_info)}...")
    
    # 批量翻译
    zh_to_en_translations = []
    en_to_zh_translations = []
    
    if zh_to_en_texts:
        print("\n执行中文→英文批量翻译...")
        zh_to_en_translations = batch_translate(
            translators['zh_to_en'], 
            zh_to_en_texts,
            batch_size=batch_size
        )
    
    if en_to_zh_texts:
        print("\n执行英文→中文批量翻译...")
        en_to_zh_translations = batch_translate(
            translators['en_to_zh'], 
            en_to_zh_texts,
            batch_size=batch_size
        )
    
    # 创建翻译结果的映射字典
    zh_to_en_map = dict(zip(zh_to_en_texts, zh_to_en_translations)) if zh_to_en_texts else {}
    en_to_zh_map = dict(zip(en_to_zh_texts, en_to_zh_translations)) if en_to_zh_texts else {}
    
    # 将翻译结果插入到DataFrame中，紧跟在原列后面
    print("\n将翻译结果添加到数据...")
    
    # 记录列的原始顺序
    original_columns = list(df.columns)
    # 创建一个新的DataFrame用于保存结果
    result_df = pd.DataFrame()
    
    # 遍历所有列，如果是需要翻译的列，则添加原列和翻译列；否则只添加原列
    for i, col in enumerate(original_columns):
        # 添加原始列
        result_df[col] = df[col]
        
        # 如果当前列需要中文→英文翻译
        if col in zh_to_en_columns:
            # 创建新列名
            new_col = f"{col}_en"
            
            # 创建新列数据
            new_col_data = [f"{df[col].iloc[0]}_en"]  # 表头添加_en后缀
            for i in range(1, len(df)):
                val = df[col].iloc[i]
                if pd.notnull(val) and isinstance(val, str):
                    new_col_data.append(zh_to_en_map.get(val, ""))
                else:
                    new_col_data.append("")
            
            # 添加翻译列
            result_df[new_col] = new_col_data
        
        # 如果当前列需要英文→中文翻译
        elif col in en_to_zh_columns:
            # 创建新列名
            new_col = f"{col}_zh"
            
            # 创建新列数据
            new_col_data = [f"{df[col].iloc[0]}_zh"]  # 表头添加_zh后缀
            for i in range(1, len(df)):
                val = df[col].iloc[i]
                if pd.notnull(val) and isinstance(val, str):
                    new_col_data.append(en_to_zh_map.get(val, ""))
                else:
                    new_col_data.append("")
            
            # 添加翻译列
            result_df[new_col] = new_col_data
    
    # 保存为新的Excel文件
    print(f"\n保存翻译结果到 {output_path}")
    result_df.to_excel(output_path, index=False)
    
    # 计算耗时
    end_time = time.time()
    elapsed_time = end_time - start_time
    elapsed_str = str(datetime.timedelta(seconds=int(elapsed_time)))
    
    print(f'\n翻译完成，已生成 {output_path}')
    print(f'总耗时: {elapsed_str} (时:分:秒)')
    
    # 输出结果总结
    if zh_to_en_indices:
        zh_to_en_cols = [index_to_column_letter(idx) for idx in zh_to_en_indices]
        print(f'- {", ".join(zh_to_en_cols)}列：中文→英文')
    if en_to_zh_indices:
        en_to_zh_cols = [index_to_column_letter(idx) for idx in en_to_zh_indices]
        print(f'- {", ".join(en_to_zh_cols)}列：英文→中文')
    
    return output_path

def translate_via_csv(input_path, translators, zh_to_en_indices, en_to_zh_indices, batch_size=10):
    """通过CSV中间格式执行Excel文件翻译，提高大文件处理效率"""
    # 记录开始时间
    start_time = time.time()
    
    # 自动生成输出文件名
    name, ext = os.path.splitext(os.path.basename(input_path))
    output_filename = f'{name}_translated{ext}'
    output_path = os.path.join(os.path.dirname(input_path), output_filename)
    temp_csv = os.path.join(os.path.dirname(input_path), f'{name}_temp.csv')
    
    print(f"\n正在加载 {os.path.basename(input_path)} 并转换为CSV...")
    
    # 使用pandas读取Excel文件
    df = pd.read_excel(input_path)
    
    # 收集所有需要翻译的文本
    zh_to_en_columns = []
    en_to_zh_columns = []
    
    # 将索引转换为列名
    for idx in zh_to_en_indices:
        zh_to_en_columns.append(df.columns[idx])
    
    for idx in en_to_zh_indices:
        en_to_zh_columns.append(df.columns[idx])
    
    # 保存需要翻译的列的原始名称，后面用于创建新列
    zh_to_en_orig_names = zh_to_en_columns.copy()
    en_to_zh_orig_names = en_to_zh_columns.copy()
    
    # 提取所有不同的文本，减少翻译量
    zh_to_en_texts = {}
    en_to_zh_texts = {}
    
    print("正在从CSV提取需要翻译的文本...")
    
    # 收集中文->英文翻译文本
    for col in zh_to_en_columns:
        # 跳过空值和非字符串值
        unique_texts = df[col].dropna().astype(str).unique()
        for text in unique_texts:
            if text and isinstance(text, str):
                zh_to_en_texts[text] = ""
    
    # 收集英文->中文翻译文本
    for col in en_to_zh_columns:
        # 跳过空值和非字符串值
        unique_texts = df[col].dropna().astype(str).unique()
        for text in unique_texts:
            if text and isinstance(text, str):
                en_to_zh_texts[text] = ""
    
    print(f"需要翻译的唯一文本: 中->英 {len(zh_to_en_texts)}个, 英->中 {len(en_to_zh_texts)}个")
    
    # 批量翻译
    if zh_to_en_texts:
        print("\n执行中文→英文批量翻译...")
        # 将字典的键转为列表以便批量翻译
        texts_to_translate = list(zh_to_en_texts.keys())
        
        # 使用线程池进行并行翻译以提高效率
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            # 分批处理文本
            batches = [texts_to_translate[i:i+batch_size] for i in range(0, len(texts_to_translate), batch_size)]
            
            # 创建翻译任务
            future_to_batch = {
                executor.submit(batch_translate, translators['zh_to_en'], batch, 1, 0.5): i 
                for i, batch in enumerate(batches)
            }
            
            # 收集结果
            batch_idx = 0
            for future in tqdm(concurrent.futures.as_completed(future_to_batch), total=len(batches), desc="翻译批次"):
                batch_idx = future_to_batch[future]
                try:
                    translations = future.result()
                    # 更新翻译结果字典
                    start_idx = batch_idx * batch_size
                    for i, translation in enumerate(translations):
                        if start_idx + i < len(texts_to_translate):
                            zh_to_en_texts[texts_to_translate[start_idx + i]] = translation
                except Exception as e:
                    print(f"批次 {batch_idx} 翻译出错: {e}")
    
    if en_to_zh_texts:
        print("\n执行英文→中文批量翻译...")
        # 将字典的键转为列表以便批量翻译
        texts_to_translate = list(en_to_zh_texts.keys())
        
        # 使用线程池进行并行翻译以提高效率
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            # 分批处理文本
            batches = [texts_to_translate[i:i+batch_size] for i in range(0, len(texts_to_translate), batch_size)]
            
            # 创建翻译任务
            future_to_batch = {
                executor.submit(batch_translate, translators['en_to_zh'], batch, 1, 0.5): i 
                for i, batch in enumerate(batches)
            }
            
            # 收集结果
            batch_idx = 0
            for future in tqdm(concurrent.futures.as_completed(future_to_batch), total=len(batches), desc="翻译批次"):
                batch_idx = future_to_batch[future]
                try:
                    translations = future.result()
                    # 更新翻译结果字典
                    start_idx = batch_idx * batch_size
                    for i, translation in enumerate(translations):
                        if start_idx + i < len(texts_to_translate):
                            en_to_zh_texts[texts_to_translate[start_idx + i]] = translation
                except Exception as e:
                    print(f"批次 {batch_idx} 翻译出错: {e}")
    
    # 将翻译结果插入到DataFrame中，紧跟在原列后面
    print("\n将翻译结果添加到数据...")
    
    # 记录列的原始顺序
    original_columns = list(df.columns)
    # 创建一个新的DataFrame用于保存结果
    result_df = pd.DataFrame()
    
    # 遍历所有列，如果是需要翻译的列，则添加原列和翻译列；否则只添加原列
    for i, col in enumerate(original_columns):
        # 添加原始列
        result_df[col] = df[col]
        
        # 如果当前列需要中文→英文翻译
        if col in zh_to_en_orig_names:
            # 创建新列名
            new_col = f"{col}_en"
            # 创建新列数据
            new_col_data = df[col].map(lambda x: zh_to_en_texts.get(str(x), "") if pd.notnull(x) and isinstance(x, str) else "")
            # 添加翻译列
            result_df[new_col] = new_col_data
        
        # 如果当前列需要英文→中文翻译
        elif col in en_to_zh_orig_names:
            # 创建新列名
            new_col = f"{col}_zh"
            # 创建新列数据
            new_col_data = df[col].map(lambda x: en_to_zh_texts.get(str(x), "") if pd.notnull(x) and isinstance(x, str) else "")
            # 添加翻译列
            result_df[new_col] = new_col_data
    
    # 保存为新的Excel文件
    print(f"\n保存翻译结果到 {output_path}")
    result_df.to_excel(output_path, index=False)
    
    # 计算耗时
    end_time = time.time()
    elapsed_time = end_time - start_time
    elapsed_str = str(datetime.timedelta(seconds=int(elapsed_time)))
    
    print(f'\n翻译完成，已生成 {output_path}')
    print(f'总耗时: {elapsed_str} (时:分:秒)')
    
    # 输出结果总结
    if zh_to_en_indices:
        zh_to_en_cols = [index_to_column_letter(idx) for idx in zh_to_en_indices]
        print(f'- {", ".join(zh_to_en_cols)}列：中文→英文')
    if en_to_zh_indices:
        en_to_zh_cols = [index_to_column_letter(idx) for idx in en_to_zh_indices]
        print(f'- {", ".join(en_to_zh_cols)}列：英文→中文')
    
    return output_path

def interactive_mode():
    """交互式模式主函数"""
    try:
        # 显示标题
        display_header()
        
        # 选择Excel文件
        input_path = get_excel_file()
        if not input_path:
            return
        
        # 加载Excel文件以显示列信息
        wb = openpyxl.load_workbook(input_path)
        
        # 选择翻译API
        translators = select_translator()
        
        # 选择要翻译的列
        zh_to_en_indices, en_to_zh_indices = select_columns(wb)
        if zh_to_en_indices is None and en_to_zh_indices is None:
            return
        
        # 设置批量翻译大小
        batch_size = get_batch_size()
        
        # 询问是否使用CSV中间格式
        use_csv = input("\n是否使用CSV中间格式加速翻译(适合大文件)？(y/n): ").strip().lower() == 'y'
        
        # 执行翻译
        if use_csv:
            output_path = translate_via_csv(
                input_path, 
                translators, 
                zh_to_en_indices, 
                en_to_zh_indices, 
                batch_size
            )
        else:
            output_path = translate_excel_file(
                input_path, 
                translators, 
                zh_to_en_indices, 
                en_to_zh_indices, 
                batch_size
            )

        print("\n感谢使用Excel自动翻译工具！")
        
        # 询问是否打开生成的文件
        open_file = input("\n是否打开生成的Excel文件？(y/n): ").strip().lower()
        if open_file == 'y':
            try:
                import platform
                import subprocess
                
                system = platform.system()
                if system == 'Darwin':  # macOS
                    subprocess.call(['open', output_path])
                elif system == 'Windows':
                    os.startfile(output_path)
                elif system == 'Linux':
                    subprocess.call(['xdg-open', output_path])
                else:
                    print(f"无法自动打开文件，请手动打开: {output_path}")
            except Exception as e:
                print(f"打开文件失败: {e}")
                print(f"请手动打开文件: {output_path}")
        
    except Exception as e:
        print(f"发生错误: {e}")
        import traceback
        print(traceback.format_exc())

def main():
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='Excel文件自动翻译工具')
    parser.add_argument('-i', '--interactive', action='store_true', help='使用交互式模式')
    parser.add_argument('-f', '--file', type=str, help='Excel文件路径')
    parser.add_argument('--zh2en', type=str, help='中文翻译成英文的列（如A,B,C等）')
    parser.add_argument('--en2zh', type=str, help='英文翻译成中文的列（如A,B,C等）')
    parser.add_argument('--api', type=int, choices=[1, 2, 3], help='翻译API选择：1=MyMemory, 2=Google, 3=百度')
    parser.add_argument('--batch', type=int, default=10, help='批量翻译大小')
    parser.add_argument('--baidu-appid', type=str, help='百度翻译API的APP ID')
    parser.add_argument('--baidu-key', type=str, help='百度翻译API的密钥')
    parser.add_argument('--use-csv', action='store_true', help='使用CSV中间格式加速翻译(适合大文件)')
    
    # 解析命令行参数
    args = parser.parse_args()
    
    # 如果指定了交互式模式或没有提供任何参数，则进入交互模式
    if args.interactive or len(sys.argv) == 1:
        interactive_mode()
    else:
        # 命令行模式
        if not args.file:
            print("错误：请指定Excel文件路径")
            parser.print_help()
            return
        
        if not args.zh2en and not args.en2zh:
            print("错误：请至少指定一个翻译方向（--zh2en或--en2zh）")
            parser.print_help()
            return
        
        # 初始化翻译器
        translators = {}
        api_choice = args.api if args.api else 1
        
        try:
            if api_choice == 1:  # MyMemory翻译
                translators['zh_to_en'] = MyMemoryTranslator(source='zh-CN', target='en-GB')
                translators['en_to_zh'] = MyMemoryTranslator(source='en-GB', target='zh-CN')
            elif api_choice == 2:  # Google翻译
                translators['zh_to_en'] = GoogleTranslator(source='zh-CN', target='en')
                translators['en_to_zh'] = GoogleTranslator(source='en', target='zh-CN')
            elif api_choice == 3:  # 百度翻译
                if not args.baidu_appid or not args.baidu_key:
                    print("错误：使用百度翻译API需要提供APP ID和密钥")
                    return
                
                translators['zh_to_en'] = BaiduTranslator(
                    appid=args.baidu_appid,
                    appkey=args.baidu_key,
                    source='zh', 
                    target='en'
                )
                translators['en_to_zh'] = BaiduTranslator(
                    appid=args.baidu_appid,
                    appkey=args.baidu_key,
                    source='en',
                    target='zh'
                )
        except Exception as e:
            print(f"初始化翻译API失败: {e}")
            return
        
        # 解析列参数
        zh_to_en_indices = []
        en_to_zh_indices = []
        
        if args.zh2en:
            columns = parse_column_input(args.zh2en)
            for col in columns:
                try:
                    zh_to_en_indices.append(column_letter_to_index(col))
                except:
                    print(f"错误：无效的列名 '{col}'")
                    return
        
        if args.en2zh:
            columns = parse_column_input(args.en2zh)
            for col in columns:
                try:
                    en_to_zh_indices.append(column_letter_to_index(col))
                except:
                    print(f"错误：无效的列名 '{col}'")
                    return
        
        # 执行翻译
        if args.use_csv:
            translate_via_csv(
                args.file, 
                translators, 
                zh_to_en_indices, 
                en_to_zh_indices, 
                args.batch
            )
        else:
            translate_excel_file(
                args.file, 
                translators, 
                zh_to_en_indices, 
                en_to_zh_indices, 
                args.batch
            )

if __name__ == "__main__":
    main() 