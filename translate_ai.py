import openpyxl
from deep_translator import GoogleTranslator, MyMemoryTranslator, BaiduTranslator
import os
from tqdm import tqdm
import string
import time
from collections import defaultdict
import sys
import argparse
import re
import pandas as pd
import csv
import concurrent.futures
import datetime
import requests
import json
# 条件导入dotenv和configparser
try:
    import dotenv
    dotenv_available = True
except ImportError:
    dotenv_available = False

try:
    import configparser
    from pathlib import Path
    config_available = True
except ImportError:
    config_available = False

def column_letter_to_index(column_letter):
    """将列字母转换为索引（A=0, B=1, ...）"""
    return string.ascii_uppercase.index(column_letter.upper())

def index_to_column_letter(index):
    """将索引转换为列字母（0=A, 1=B, ...）"""
    return string.ascii_uppercase[index]

# 批量翻译函数
def batch_translate(translator, texts, batch_size=10, delay=1):
    """批量翻译文本，减少API调用次数"""
    if not texts:
        return []
        
    # 去重以减少翻译量
    unique_texts = list(set(texts))
    print(f"需要翻译 {len(texts)} 个单元格，去重后 {len(unique_texts)} 个唯一文本")
    
    # 创建翻译缓存
    translation_cache = {}
    results = []
    
    # 检查翻译器类型，为不同翻译器采用不同策略
    is_deepseek = isinstance(translator, DeepSeekTranslator)
    
    # 如果是DeepSeek-V3，使用更高效的批处理方式
    if is_deepseek:
        # DeepSeek可以处理更大的文本，合并更多文本减少API调用
        max_chars = 3500  # 设置每批次最大字符数
        current_batch = []
        current_char_count = 0
        
        print("使用DeepSeek-V3进行高效批量翻译...")
        
        # 基于字符计数而不是固定批次大小来分批
        batches = []
        for text in unique_texts:
            text_len = len(text)
            # 如果单条文本就超过限制，单独处理
            if text_len > max_chars:
                batches.append([text])
                continue
                
            # 如果添加当前文本会超出限制，创建新批次
            if current_char_count + text_len + len(current_batch) * 5 > max_chars and current_batch:  # 5是分隔符长度
                batches.append(current_batch)
                current_batch = [text]
                current_char_count = text_len
            else:
                current_batch.append(text)
                current_char_count += text_len
        
        # 添加最后一个批次
        if current_batch:
            batches.append(current_batch)
            
        print(f"DeepSeek-V3批处理：将{len(unique_texts)}个文本分为{len(batches)}个批次翻译")
        
        for batch in tqdm(batches, desc="DeepSeek翻译进度"):
            try:
                # 使用特殊分隔符合并文本
                combined_text = " [SEP] ".join(batch)
                translated = translator.translate(combined_text)
                
                # 尝试拆分翻译结果
                translated_parts = translated.split(" [SEP] ")
                
                # 如果拆分结果与原文数量不符，尝试其他分隔方式
                if len(translated_parts) != len(batch):
                    # 尝试识别分隔符的其他可能变体
                    possible_seps = [" [SEP] ", "[SEP]", " [sep] ", "[sep]", " ; ", ";", "。", ". "]
                    for sep in possible_seps:
                        translated_parts = translated.split(sep)
                        if len(translated_parts) == len(batch):
                            break
                
                # 如果仍然无法正确拆分，逐个翻译
                if len(translated_parts) != len(batch):
                    print(f"\n批量拆分异常，切换为逐个翻译...")
                    for text in batch:
                        trans = translator.translate(text)
                        translation_cache[text] = trans
                else:
                    # 缓存翻译结果
                    for i, text in enumerate(batch):
                        if i < len(translated_parts):
                            translation_cache[text] = translated_parts[i]
                        else:
                            # 处理索引越界情况
                            translation_cache[text] = text
                            
                # DeepSeek API可能有速率限制，添加短暂延迟
                if delay > 0:
                    time.sleep(delay)
                    
            except Exception as e:
                print(f"\n批次翻译失败: {str(e)}，切换为逐个翻译...")
                for text in batch:
                    try:
                        trans = translator.translate(text)
                        translation_cache[text] = trans
                        time.sleep(0.5)  # 单条翻译添加更短的延迟
                    except Exception as e:
                        print(f"单条翻译失败: {text[:30]}..., 错误: {str(e)}")
                        translation_cache[text] = text  # 失败时用原文
    else:
        # 将文本分批处理
        batches = [unique_texts[i:i+batch_size] for i in range(0, len(unique_texts), batch_size)]
        
        for i, batch in enumerate(tqdm(batches, desc="批次进度")):
            try:
                # 将多个文本合并为一个长文本，用特殊分隔符隔开
                combined_text = " ||| ".join(batch)
                
                # 翻译合并后的文本
                translated = translator.translate(combined_text)
                
                # 根据同样的分隔符拆分翻译结果
                translated_parts = translated.split(" ||| ")
                
                # 如果翻译结果数量与原文本不符，则逐个翻译
                if len(translated_parts) != len(batch):
                    print(f"\n批量翻译结果异常，切换为逐个翻译...")
                    for text in batch:
                        try:
                            trans = translator.translate(text)
                            translation_cache[text] = trans
                        except Exception as e:
                            print(f"翻译失败: {text}, 错误: {str(e)}")
                            translation_cache[text] = text  # 失败时用原文
                else:
                    # 缓存翻译结果
                    for j, text in enumerate(batch):
                        translation_cache[text] = translated_parts[j]
                
                # 批次之间添加延迟，避免API限制
                if i < len(batches) - 1 and delay > 0:
                    time.sleep(delay)
                    
            except Exception as e:
                print(f"\n批次翻译失败，切换为逐个翻译...")
                for text in batch:
                    try:
                        trans = translator.translate(text)
                        translation_cache[text] = trans
                    except Exception as e:
                        print(f"翻译失败: {text}, 错误: {str(e)}")
                        translation_cache[text] = text  # 失败时用原文
    
    # 根据原始顺序返回翻译结果
    for text in texts:
        results.append(translation_cache.get(text, text))
    
    return results

def display_header():
    """显示应用程序标题"""
    print("="*60)
    print("Excel自动翻译工具 - 支持中英文互译")
    print("="*60)

def get_excel_file():
    """交互式选择Excel文件"""
    # 获取当前目录下所有Excel文件
    excel_files = [f for f in os.listdir('.') if f.endswith(('.xlsx', '.xls'))]
    
    if not excel_files:
        print("当前目录下没有找到Excel文件(.xlsx或.xls)")
        file_path = input("请输入Excel文件的完整路径: ")
        if not file_path.endswith(('.xlsx', '.xls')):
            file_path += '.xlsx'
    else:
        # 显示可选的Excel文件
        print("\n当前目录下的Excel文件:")
        for i, file in enumerate(excel_files, 1):
            print(f"{i}. {file}")
        
        # 让用户选择文件
        choice = input("\n请选择要翻译的Excel文件序号（或输入其他文件路径）: ")
        try:
            index = int(choice) - 1
            if 0 <= index < len(excel_files):
                file_path = excel_files[index]
            else:
                print("无效的选择，请输入有效的序号或文件路径")
                return None
        except ValueError:
            # 用户可能直接输入了文件路径
            file_path = choice
            if not file_path.endswith(('.xlsx', '.xls')):
                file_path += '.xlsx'
    
    # 检查文件是否存在
    if not os.path.exists(file_path):
        print(f"错误：找不到文件 {file_path}。请确认文件名并重试。")
        return None
        
    return file_path

# 加载环境变量和配置
def load_config():
    """加载配置，包括环境变量和配置文件"""
    config = {
        'deepseek_key': 'sk-ebb7d2c79ec24a63951186dba80e8ee0',
        'deepseek_url': 'https://api.deepseek.com',
        'baidu_appid': None,
        'baidu_key': None
    }
    
    # 1. 尝试加载.env文件
    if dotenv_available:
        try:
            dotenv.load_dotenv()
        except Exception as e:
            print(f"加载.env文件时出错: {e}")
    
    # 2. 从环境变量读取配置
    config['deepseek_key'] = os.environ.get('DEEPSEEK_API_KEY')
    config['deepseek_url'] = os.environ.get('DEEPSEEK_API_URL')
    config['baidu_appid'] = os.environ.get('BAIDU_API_ID')
    config['baidu_key'] = os.environ.get('BAIDU_API_KEY')
    
    # 3. 尝试读取配置文件
    if config_available:
        try:
            config_paths = [
                Path('config.ini'),
                Path('~/.excel_translator/config.ini').expanduser(),
                Path(os.path.dirname(os.path.abspath(__file__))) / 'config.ini'
            ]
            
            config_parser = configparser.ConfigParser()
            for path in config_paths:
                if path.exists():
                    try:
                        config_parser.read(path)
                        if 'api' in config_parser:
                            api_section = config_parser['api']
                            if not config['deepseek_key'] and 'deepseek_key' in api_section:
                                config['deepseek_key'] = api_section.get('deepseek_key')
                            if not config['deepseek_url'] and 'deepseek_url' in api_section:
                                config['deepseek_url'] = api_section.get('deepseek_url')
                            if not config['baidu_appid'] and 'baidu_appid' in api_section:
                                config['baidu_appid'] = api_section.get('baidu_appid')
                            if not config['baidu_key'] and 'baidu_key' in api_section:
                                config['baidu_key'] = api_section.get('baidu_key')
                        break
                    except Exception as e:
                        print(f"读取配置文件 {path} 出错: {e}")
        except Exception as e:
            print(f"处理配置文件时出错: {e}")
    
    return config

def select_translator():
    """交互式选择翻译API"""
    print("\n请选择翻译API（某些地区无法使用Google翻译）:")
    print("1. MyMemory翻译 (免费，无需密钥，推荐首选)")
    print("2. Google翻译 (部分地区可能无法访问)")
    print("3. 百度翻译 (需要API ID和密钥)")
    print("4. DeepSeek-V3 (需要API密钥)")

    translators = {}
    
    # 加载配置文件和环境变量中的API密钥
    config = load_config()
    
    while True:
        api_choice = input("请选择翻译API (1-4): ").strip()
        
        try:
            if api_choice == '1':  # MyMemory翻译
                # MyMemory对于中文使用"zh-CN"，对于英文使用"en-GB"
                translators['zh_to_en'] = MyMemoryTranslator(source='zh-CN', target='en-GB')
                translators['en_to_zh'] = MyMemoryTranslator(source='en-GB', target='zh-CN')
                break
                
            elif api_choice == '2':  # Google翻译
                translators['zh_to_en'] = GoogleTranslator(source='zh-CN', target='en')
                translators['en_to_zh'] = GoogleTranslator(source='en', target='zh-CN')
                break
                
            elif api_choice == '3':  # 百度翻译
                # 尝试从配置中获取，如果没有则提示输入
                baidu_appid = config['baidu_appid']
                baidu_key = config['baidu_key']
                
                if not baidu_appid:
                    baidu_appid = input("请输入百度翻译API的APP ID: ")
                else:
                    print(f"使用配置中的百度翻译APP ID: {baidu_appid[:4]}***")
                    override = input("是否使用其他APP ID？(y/n，默认n): ").strip().lower()
                    if override == 'y':
                        baidu_appid = input("请输入百度翻译API的APP ID: ")
                        
                if not baidu_key:
                    baidu_key = input("请输入百度翻译API的密钥: ")
                else:
                    print(f"使用配置中的百度翻译密钥: {baidu_key[:4]}***")
                    override = input("是否使用其他密钥？(y/n，默认n): ").strip().lower()
                    if override == 'y':
                        baidu_key = input("请输入百度翻译API的密钥: ")
                
                try:
                    translators['zh_to_en'] = BaiduTranslator(
                        appid=baidu_appid,
                        appkey=baidu_key,
                        source='zh', 
                        target='en'
                    )
                    translators['en_to_zh'] = BaiduTranslator(
                        appid=baidu_appid,
                        appkey=baidu_key,
                        source='en',
                        target='zh'
                    )
                    break
                except Exception as e:
                    print(f"初始化百度翻译API失败: {e}")
                    continue
            elif api_choice == '4':  # DeepSeek-V3
                # 尝试从配置中获取，如果没有则提示输入
                deepseek_key = config['deepseek_key']
                deepseek_url = config['deepseek_url']
                
                if not deepseek_key:
                    deepseek_key = input("请输入 DeepSeek-V3 API 密钥: ")
                else:
                    print(f"使用配置中的DeepSeek-V3密钥: {deepseek_key[:4]}***")
                    override = input("是否使用其他密钥？(y/n，默认n): ").strip().lower()
                    if override == 'y':
                        deepseek_key = input("请输入 DeepSeek-V3 API 密钥: ")
                
                if not deepseek_url:
                    custom_api_url = input("请输入 API URL (回车使用默认): ").strip()
                    deepseek_url = custom_api_url if custom_api_url else None
                else:
                    print(f"使用配置中的DeepSeek-V3 URL: {deepseek_url}")
                    override = input("是否使用其他URL？(y/n，默认n): ").strip().lower()
                    if override == 'y':
                        custom_api_url = input("请输入 API URL (回车使用默认): ").strip()
                        deepseek_url = custom_api_url if custom_api_url else deepseek_url
                
                try:
                    translators['zh_to_en'] = DeepSeekTranslator(
                        source='zh-CN', 
                        target='en',
                        api_key=deepseek_key,
                        api_url=deepseek_url
                    )
                    translators['en_to_zh'] = DeepSeekTranslator(
                        source='en', 
                        target='zh-CN',
                        api_key=deepseek_key,
                        api_url=deepseek_url
                    )
                    break
                except Exception as e:
                    print(f"初始化 DeepSeek-V3 API 失败: {e}")
                    continue
            else:
                print("无效选择，请重新输入")
                
        except Exception as e:
            print(f"初始化翻译API失败: {e}")
            print("请重新选择翻译API")
    
    return translators

def parse_column_input(column_input):
    """解析用户输入的列名，支持多种分隔符和格式"""
    if not column_input:
        return []
    
    # 移除所有空格
    column_input = column_input.replace(' ', '')
    
    # 使用正则表达式匹配列名（支持A、B、C或A,B,C或A，B，C等格式）
    columns = re.findall(r'[A-Za-z]', column_input)
    return [col.upper() for col in columns]

def select_columns(wb):
    """交互式选择要翻译的列，支持多列选择"""
    ws = wb.active
    max_col = ws.max_column
    
    # 显示表头信息
    print("\n文件中的列信息:")
    headers = {}
    for col_idx in range(1, max_col + 1):
        col_letter = index_to_column_letter(col_idx - 1)
        header_value = ws.cell(row=1, column=col_idx).value
        headers[col_letter] = header_value
        print(f"{col_letter}. {header_value}")
    
    print("\n请指定要翻译的列：(如不需要某方向翻译，请直接按回车跳过)")
    print("支持多列输入，例如: A,B,C 或 A、B、C 或 ABC")
    
    zh_to_en_input = input("哪些列需要从【中文翻译成英文】（输入列字母，如A、B等，不需要则直接按回车）: ").strip()
    en_to_zh_input = input("哪些列需要从【英文翻译成中文】（输入列字母，如A、B等，不需要则直接按回车）: ").strip()
    
    # 解析用户输入的列名
    zh_to_en_columns = parse_column_input(zh_to_en_input)
    en_to_zh_columns = parse_column_input(en_to_zh_input)
    
    # 验证至少有一个翻译方向
    if not zh_to_en_columns and not en_to_zh_columns:
        print("错误：请至少指定一个翻译方向")
        return None, None
    
    # 验证列名有效性
    valid_columns = list(string.ascii_uppercase[:26])  # 只支持A-Z
    zh_to_en_indices = []
    en_to_zh_indices = []
    
    # 验证中文→英文的列
    for col in zh_to_en_columns:
        if col not in valid_columns:
            print(f"错误：'{col}'不是有效的列名（A-Z）")
            return None, None
        col_idx = column_letter_to_index(col)
        if col_idx >= max_col:
            print(f"错误：列{col}超出文件范围，文件只有{max_col}列")
            return None, None
        zh_to_en_indices.append(col_idx)
    
    # 验证英文→中文的列
    for col in en_to_zh_columns:
        if col not in valid_columns:
            print(f"错误：'{col}'不是有效的列名（A-Z）")
            return None, None
        col_idx = column_letter_to_index(col)
        if col_idx >= max_col:
            print(f"错误：列{col}超出文件范围，文件只有{max_col}列")
            return None, None
        en_to_zh_indices.append(col_idx)
    
    # 检查是否有重复选择的列
    duplicate_cols = set(zh_to_en_columns) & set(en_to_zh_columns)
    if duplicate_cols:
        print(f"警告：列 {', '.join(duplicate_cols)} 同时被选择为中文→英文和英文→中文翻译")
        confirm_duplicate = input("是否继续？(y/n): ").strip().lower()
        if confirm_duplicate != 'y':
            return None, None
    
    # 输出确认信息
    print(f"\n将翻译：")
    if zh_to_en_columns:
        print("从中文翻译成英文的列:")
        for col in zh_to_en_columns:
            print(f"- {col}列 ({headers.get(col, '未知表头')})")
    
    if en_to_zh_columns:
        print("从英文翻译成中文的列:")
        for col in en_to_zh_columns:
            print(f"- {col}列 ({headers.get(col, '未知表头')})")
    
    confirm = input("\n请确认(y/n): ").strip().lower()
    if confirm != 'y':
        print("已取消操作")
        return None, None
    
    return zh_to_en_indices, en_to_zh_indices

def get_batch_size():
    """交互式设置批量翻译大小"""
    default_batch_size = 10
    try:
        custom_batch = input(f"\n设置批量翻译大小（默认每批{default_batch_size}个，较大的值翻译更快但可能失败）: ").strip()
        if custom_batch:
            batch_size = int(custom_batch)
        else:
            batch_size = default_batch_size
    except:
        print(f"使用默认批量大小: {default_batch_size}")
        batch_size = default_batch_size
    
    return batch_size

def translate_excel_file(input_path, translators, zh_to_en_indices, en_to_zh_indices, batch_size=10):
    """执行Excel文件翻译，支持多列翻译"""
    # 记录开始时间
    start_time = time.time()
    
    # 自动生成输出文件名
    name, ext = os.path.splitext(os.path.basename(input_path))
    output_filename = f'{name}_translated{ext}'
    output_path = os.path.join(os.path.dirname(input_path), output_filename)
    
    # 加载Excel文件
    print(f"\n正在加载 {os.path.basename(input_path)}...")
    
    # 使用pandas读取Excel文件，而不是openpyxl，可以更方便地处理列的插入
    df = pd.read_excel(input_path)
    max_row, max_col = df.shape
    
    print(f"文件加载完成，共有 {max_row} 行，{max_col} 列")
    
    # 收集所有需要翻译的文本
    zh_to_en_columns = []
    en_to_zh_columns = []
    zh_to_en_texts = []
    en_to_zh_texts = []
    
    # 将索引转换为列名
    for idx in zh_to_en_indices:
        zh_to_en_columns.append(df.columns[idx])
    
    for idx in en_to_zh_indices:
        en_to_zh_columns.append(df.columns[idx])
    
    # 收集要翻译的文本
    print("正在收集需要翻译的文本...")
    for col in zh_to_en_columns:
        # 跳过表头，只翻译内容
        texts = df[col].iloc[1:].dropna().astype(str).tolist()
        zh_to_en_texts.extend(texts)
    
    for col in en_to_zh_columns:
        # 跳过表头，只翻译内容
        texts = df[col].iloc[1:].dropna().astype(str).tolist()
        en_to_zh_texts.extend(texts)
    
    # 打印开始翻译的信息
    translate_info = []
    if zh_to_en_indices:
        zh_to_en_cols = [index_to_column_letter(idx) for idx in zh_to_en_indices]
        translate_info.append(f"{','.join(zh_to_en_cols)}列(中→英)")
    if en_to_zh_indices:
        en_to_zh_cols = [index_to_column_letter(idx) for idx in en_to_zh_indices]
        translate_info.append(f"{','.join(en_to_zh_cols)}列(英→中)")
    print(f"\n开始翻译{' 和 '.join(translate_info)}...")
    
    # 批量翻译
    zh_to_en_translations = []
    en_to_zh_translations = []
    
    if zh_to_en_texts:
        print("\n执行中文→英文批量翻译...")
        zh_to_en_translations = batch_translate(
            translators['zh_to_en'], 
            zh_to_en_texts,
            batch_size=batch_size
        )
    
    if en_to_zh_texts:
        print("\n执行英文→中文批量翻译...")
        en_to_zh_translations = batch_translate(
            translators['en_to_zh'], 
            en_to_zh_texts,
            batch_size=batch_size
        )
    
    # 创建翻译结果的映射字典
    zh_to_en_map = dict(zip(zh_to_en_texts, zh_to_en_translations)) if zh_to_en_texts else {}
    en_to_zh_map = dict(zip(en_to_zh_texts, en_to_zh_translations)) if en_to_zh_texts else {}
    
    # 将翻译结果插入到DataFrame中，紧跟在原列后面
    print("\n将翻译结果添加到数据...")
    
    # 记录列的原始顺序
    original_columns = list(df.columns)
    # 创建一个新的DataFrame用于保存结果
    result_df = pd.DataFrame()
    
    # 遍历所有列，如果是需要翻译的列，则添加原列和翻译列；否则只添加原列
    for i, col in enumerate(original_columns):
        # 添加原始列
        result_df[col] = df[col]
        
        # 如果当前列需要中文→英文翻译
        if col in zh_to_en_columns:
            # 创建新列名
            new_col = f"{col}_en"
            
            # 创建新列数据
            new_col_data = [f"{df[col].iloc[0]}_en"]  # 表头添加_en后缀
            for i in range(1, len(df)):
                val = df[col].iloc[i]
                if pd.notnull(val) and isinstance(val, str):
                    new_col_data.append(zh_to_en_map.get(val, ""))
                else:
                    new_col_data.append("")
            
            # 添加翻译列
            result_df[new_col] = new_col_data
        
        # 如果当前列需要英文→中文翻译
        elif col in en_to_zh_columns:
            # 创建新列名
            new_col = f"{col}_zh"
            
            # 创建新列数据
            new_col_data = [f"{df[col].iloc[0]}_zh"]  # 表头添加_zh后缀
            for i in range(1, len(df)):
                val = df[col].iloc[i]
                if pd.notnull(val) and isinstance(val, str):
                    new_col_data.append(en_to_zh_map.get(val, ""))
                else:
                    new_col_data.append("")
            
            # 添加翻译列
            result_df[new_col] = new_col_data
    
    # 保存为新的Excel文件
    print(f"\n保存翻译结果到 {output_path}")
    result_df.to_excel(output_path, index=False)
    
    # 计算耗时
    end_time = time.time()
    elapsed_time = end_time - start_time
    elapsed_str = str(datetime.timedelta(seconds=int(elapsed_time)))
    
    print(f'\n翻译完成，已生成 {output_path}')
    print(f'总耗时: {elapsed_str} (时:分:秒)')
    
    # 输出结果总结
    if zh_to_en_indices:
        zh_to_en_cols = [index_to_column_letter(idx) for idx in zh_to_en_indices]
        print(f'- {", ".join(zh_to_en_cols)}列：中文→英文')
    if en_to_zh_indices:
        en_to_zh_cols = [index_to_column_letter(idx) for idx in en_to_zh_indices]
        print(f'- {", ".join(en_to_zh_cols)}列：英文→中文')
    
    return output_path

def translate_via_csv(input_path, translators, zh_to_en_indices, en_to_zh_indices, batch_size=10):
    """通过CSV中间格式执行Excel文件翻译，提高大文件处理效率"""
    # 记录开始时间
    start_time = time.time()
    
    # 自动生成输出文件名和临时CSV文件名
    name, ext = os.path.splitext(os.path.basename(input_path))
    output_filename = f'{name}_translated{ext}'
    output_path = os.path.join(os.path.dirname(input_path), output_filename)
    temp_csv = os.path.join(os.path.dirname(input_path), f'{name}_temp.csv')
    
    print(f"\n正在加载 {os.path.basename(input_path)} 并转换为CSV...")
    
    # 优化：直接使用pandas读取Excel文件，避免内存占用
    try:
        df = pd.read_excel(input_path)
        df.to_csv(temp_csv, index=False, encoding='utf-8-sig')
        print(f"已将Excel转换为临时CSV文件，共 {len(df)} 行数据")
    except Exception as e:
        print(f"Excel转CSV出错: {e}")
        # 尝试使用低内存方式读取
        print("尝试使用低内存模式读取...")
        df = pd.read_excel(input_path, engine='openpyxl')
        df.to_csv(temp_csv, index=False, encoding='utf-8-sig')
        
    # 收集所有需要翻译的文本，使用CSV文件读取以节省内存
    print("正在从CSV提取需要翻译的文本...")
    
    # 将索引转换为列名
    df_columns = list(df.columns)
    zh_to_en_columns = [df_columns[idx] for idx in zh_to_en_indices] if zh_to_en_indices else []
    en_to_zh_columns = [df_columns[idx] for idx in en_to_zh_indices] if en_to_zh_indices else []
    
    # 提取需要翻译的唯一文本
    zh_to_en_texts = {}
    en_to_zh_texts = {}
    
    # 使用CSV文件迭代器逐行读取，减少内存消耗
    with open(temp_csv, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in tqdm(reader, desc="提取文本"):
            # 提取中文->英文需要翻译的文本
            for col in zh_to_en_columns:
                if col in row and row[col] and isinstance(row[col], str):
                    zh_to_en_texts[row[col]] = ""
            
            # 提取英文->中文需要翻译的文本
            for col in en_to_zh_columns:
                if col in row and row[col] and isinstance(row[col], str):
                    en_to_zh_texts[row[col]] = ""
    
    # 统计需要翻译的文本数量
    print(f"需要翻译的唯一文本: 中->英 {len(zh_to_en_texts)}个, 英->中 {len(en_to_zh_texts)}个")
    
    # 检查翻译器类型
    is_deepseek = any(isinstance(t, DeepSeekTranslator) for t in translators.values())
    
    # 批量翻译中文->英文文本
    if zh_to_en_texts:
        print("\n执行中文→英文批量翻译...")
        zh_to_en_list = list(zh_to_en_texts.keys())
        
        # DeepSeek可以一次处理更多文本，提高效率
        if is_deepseek:
            translations = batch_translate(translators['zh_to_en'], zh_to_en_list, batch_size)
            for i, text in enumerate(zh_to_en_list):
                if i < len(translations):
                    zh_to_en_texts[text] = translations[i]
        else:
            # 使用线程池加速翻译，但对于DeepSeek可能会导致API限制
            batches = [zh_to_en_list[i:i+batch_size] for i in range(0, len(zh_to_en_list), batch_size)]
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                future_to_batch = {
                    executor.submit(batch_translate, translators['zh_to_en'], batch, 1, 1): i 
                    for i, batch in enumerate(batches)
                }
                
                for future in tqdm(concurrent.futures.as_completed(future_to_batch), total=len(batches), desc="翻译批次"):
                    batch_idx = future_to_batch[future]
                    try:
                        translations = future.result()
                        start_idx = batch_idx * batch_size
                        for i, translation in enumerate(translations):
                            if start_idx + i < len(zh_to_en_list):
                                zh_to_en_texts[zh_to_en_list[start_idx + i]] = translation
                    except Exception as e:
                        print(f"批次 {batch_idx} 翻译出错: {e}")
    
    # 批量翻译英文->中文文本
    if en_to_zh_texts:
        print("\n执行英文→中文批量翻译...")
        en_to_zh_list = list(en_to_zh_texts.keys())
        
        # DeepSeek可以一次处理更多文本，提高效率
        if is_deepseek:
            translations = batch_translate(translators['en_to_zh'], en_to_zh_list, batch_size)
            for i, text in enumerate(en_to_zh_list):
                if i < len(translations):
                    en_to_zh_texts[text] = translations[i]
        else:
            # 使用线程池加速翻译，但对于DeepSeek可能会导致API限制
            batches = [en_to_zh_list[i:i+batch_size] for i in range(0, len(en_to_zh_list), batch_size)]
            
            with concurrent.futures.ThreadPoolExecutor(max_workers=3) as executor:
                future_to_batch = {
                    executor.submit(batch_translate, translators['en_to_zh'], batch, 1, 1): i 
                    for i, batch in enumerate(batches)
                }
                
                for future in tqdm(concurrent.futures.as_completed(future_to_batch), total=len(batches), desc="翻译批次"):
                    batch_idx = future_to_batch[future]
                    try:
                        translations = future.result()
                        start_idx = batch_idx * batch_size
                        for i, translation in enumerate(translations):
                            if start_idx + i < len(en_to_zh_list):
                                en_to_zh_texts[en_to_zh_list[start_idx + i]] = translation
                    except Exception as e:
                        print(f"批次 {batch_idx} 翻译出错: {e}")
    
    # 逐行处理CSV并应用翻译结果
    print("\n将翻译结果应用到CSV数据...")
    
    # 准备新的CSV文件
    translated_csv = os.path.join(os.path.dirname(input_path), f'{name}_translated.csv')
    
    with open(temp_csv, 'r', encoding='utf-8-sig') as input_file, \
         open(translated_csv, 'w', encoding='utf-8-sig', newline='') as output_file:
        
        reader = csv.DictReader(input_file)
        all_columns = reader.fieldnames.copy()
        
        # 添加新列：为每个翻译列添加对应的结果列
        for col in zh_to_en_columns:
            all_columns.append(f"{col}_en")
        for col in en_to_zh_columns:
            all_columns.append(f"{col}_zh")
        
        writer = csv.DictWriter(output_file, fieldnames=all_columns)
        writer.writeheader()
        
        # 逐行处理数据
        for row in tqdm(reader, desc="生成结果"):
            new_row = {key: value for key, value in row.items()}
            
            # 添加中文->英文翻译结果
            for col in zh_to_en_columns:
                if col in row and row[col] and isinstance(row[col], str):
                    new_row[f"{col}_en"] = zh_to_en_texts.get(row[col], "")
                else:
                    new_row[f"{col}_en"] = ""
            
            # 添加英文->中文翻译结果
            for col in en_to_zh_columns:
                if col in row and row[col] and isinstance(row[col], str):
                    new_row[f"{col}_zh"] = en_to_zh_texts.get(row[col], "")
                else:
                    new_row[f"{col}_zh"] = ""
            
            writer.writerow(new_row)
    
    # 将CSV结果转回Excel格式
    print("\n将最终结果转换回Excel格式...")
    result_df = pd.read_csv(translated_csv, encoding='utf-8-sig')
    result_df.to_excel(output_path, index=False)
    
    # 清理临时文件
    try:
        os.remove(temp_csv)
        os.remove(translated_csv)
        print("已清理临时CSV文件")
    except:
        print("注意：无法删除临时CSV文件")
    
    # 计算耗时
    end_time = time.time()
    elapsed_time = end_time - start_time
    elapsed_str = str(datetime.timedelta(seconds=int(elapsed_time)))
    
    print(f'\n翻译完成，已生成 {output_path}')
    print(f'总耗时: {elapsed_str} (时:分:秒)')
    
    # 输出结果总结
    if zh_to_en_indices:
        zh_to_en_cols = [index_to_column_letter(idx) for idx in zh_to_en_indices]
        print(f'- {", ".join(zh_to_en_cols)}列：中文→英文')
    if en_to_zh_indices:
        en_to_zh_cols = [index_to_column_letter(idx) for idx in en_to_zh_indices]
        print(f'- {", ".join(en_to_zh_cols)}列：英文→中文')
    
    return output_path

def interactive_mode():
    """交互式模式主函数"""
    try:
        # 显示标题
        display_header()
        
        # 选择Excel文件
        input_path = get_excel_file()
        if not input_path:
            return
        
        # 加载Excel文件以显示列信息
        wb = openpyxl.load_workbook(input_path)
        
        # 选择翻译API
        translators = select_translator()
        
        # 选择要翻译的列
        zh_to_en_indices, en_to_zh_indices = select_columns(wb)
        if zh_to_en_indices is None and en_to_zh_indices is None:
            return
        
        # 设置批量翻译大小
        batch_size = get_batch_size()
        
        # 询问是否使用CSV中间格式
        use_csv = input("\n是否使用CSV中间格式加速翻译(适合大文件)？(y/n): ").strip().lower() == 'y'
        
        # 执行翻译
        if use_csv:
            output_path = translate_via_csv(
                input_path, 
                translators, 
                zh_to_en_indices, 
                en_to_zh_indices, 
                batch_size
            )
        else:
            output_path = translate_excel_file(
                input_path, 
                translators, 
                zh_to_en_indices, 
                en_to_zh_indices, 
                batch_size
            )

        print("\n感谢使用Excel自动翻译工具！")
        
        # 询问是否打开生成的文件
        open_file = input("\n是否打开生成的Excel文件？(y/n): ").strip().lower()
        if open_file == 'y':
            try:
                import platform
                import subprocess
                
                system = platform.system()
                if system == 'Darwin':  # macOS
                    subprocess.call(['open', output_path])
                elif system == 'Windows':
                    os.startfile(output_path)
                elif system == 'Linux':
                    subprocess.call(['xdg-open', output_path])
                else:
                    print(f"无法自动打开文件，请手动打开: {output_path}")
            except Exception as e:
                print(f"打开文件失败: {e}")
                print(f"请手动打开文件: {output_path}")
        
    except Exception as e:
        print(f"发生错误: {e}")
        import traceback
        print(traceback.format_exc())

def main():
    # 创建命令行参数解析器
    parser = argparse.ArgumentParser(description='Excel文件自动翻译工具')
    parser.add_argument('-i', '--interactive', action='store_true', help='使用交互式模式')
    parser.add_argument('-f', '--file', type=str, help='Excel文件路径')
    parser.add_argument('--zh2en', type=str, help='中文翻译成英文的列（如A,B,C等）')
    parser.add_argument('--en2zh', type=str, help='英文翻译成中文的列（如A,B,C等）')
    parser.add_argument('--api', type=int, choices=[1, 2, 3, 4], help='翻译API选择：1=MyMemory, 2=Google, 3=百度, 4=DeepSeek-V3')
    parser.add_argument('--batch', type=int, default=10, help='批量翻译大小')
    parser.add_argument('--baidu-appid', type=str, help='百度翻译API的APP ID')
    parser.add_argument('--baidu-key', type=str, help='百度翻译API的密钥')
    parser.add_argument('--deepseek-key', type=str, help='DeepSeek-V3 API的密钥')
    parser.add_argument('--deepseek-url', type=str, help='DeepSeek-V3 API的URL地址')
    parser.add_argument('--use-csv', action='store_true', help='使用CSV中间格式加速翻译(适合大文件)')
    parser.add_argument('--gen-config', action='store_true', help='生成配置文件模板')
    
    # 解析命令行参数
    args = parser.parse_args()
    
    # 如果请求生成配置文件
    if args.gen_config:
        if not config_available:
            print("错误：缺少configparser库，无法生成配置文件")
            print("请执行 'pip install configparser' 安装所需库")
            return
            
        config_path = Path('config.ini')
        if config_path.exists():
            overwrite = input(f"配置文件{config_path}已存在，是否覆盖？(y/n): ").strip().lower()
            if overwrite != 'y':
                print("取消生成配置文件")
                return
        
        try:
            config = configparser.ConfigParser()
            config['api'] = {
                'deepseek_key': 'YOUR_DEEPSEEK_API_KEY',
                'deepseek_url': 'https://api.deepseek.com/v1/chat/completions',
                'baidu_appid': 'YOUR_BAIDU_API_ID',
                'baidu_key': 'YOUR_BAIDU_API_KEY'
            }
            
            with open(config_path, 'w') as f:
                config.write(f)
            
            print(f"配置文件模板已生成: {config_path}")
            print("请编辑该文件，填入您的API密钥和URL")
            return
        except Exception as e:
            print(f"生成配置文件失败: {e}")
            return
    
    # 获取配置
    config = load_config()
    
    # 如果指定了交互式模式或没有提供任何参数，则进入交互模式
    if args.interactive or len(sys.argv) == 1:
        interactive_mode()
    else:
        # 命令行模式
        if not args.file:
            print("错误：请指定Excel文件路径")
            parser.print_help()
            return
        
        if not args.zh2en and not args.en2zh:
            print("错误：请至少指定一个翻译方向（--zh2en或--en2zh）")
            parser.print_help()
            return
        
        # 初始化翻译器
        translators = {}
        api_choice = args.api if args.api else 1
        
        try:
            if api_choice == 1:  # MyMemory翻译
                translators['zh_to_en'] = MyMemoryTranslator(source='zh-CN', target='en-GB')
                translators['en_to_zh'] = MyMemoryTranslator(source='en-GB', target='zh-CN')
            elif api_choice == 2:  # Google翻译
                translators['zh_to_en'] = GoogleTranslator(source='zh-CN', target='en')
                translators['en_to_zh'] = GoogleTranslator(source='en', target='zh-CN')
            elif api_choice == 3:  # 百度翻译
                # 优先使用命令行参数，其次使用配置
                baidu_appid = args.baidu_appid or config['baidu_appid']
                baidu_key = args.baidu_key or config['baidu_key']
                
                if not baidu_appid or not baidu_key:
                    print("错误：使用百度翻译API需要提供APP ID和密钥")
                    print("您可以通过命令行参数、环境变量或配置文件提供")
                    return
                
                translators['zh_to_en'] = BaiduTranslator(
                    appid=baidu_appid,
                    appkey=baidu_key,
                    source='zh', 
                    target='en'
                )
                translators['en_to_zh'] = BaiduTranslator(
                    appid=baidu_appid,
                    appkey=baidu_key,
                    source='en',
                    target='zh'
                )
            elif api_choice == 4:  # DeepSeek-V3
                # 优先使用命令行参数，其次使用配置
                deepseek_key = args.deepseek_key or config['deepseek_key']
                deepseek_url = args.deepseek_url or config['deepseek_url']
                
                if not deepseek_key:
                    print("错误：使用DeepSeek-V3需要提供API密钥")
                    print("您可以通过命令行参数、环境变量或配置文件提供")
                    return
                
                translators['zh_to_en'] = DeepSeekTranslator(
                    source='zh-CN', 
                    target='en',
                    api_key=deepseek_key,
                    api_url=deepseek_url
                )
                translators['en_to_zh'] = DeepSeekTranslator(
                    source='en', 
                    target='zh-CN',
                    api_key=deepseek_key,
                    api_url=deepseek_url
                )
        except Exception as e:
            print(f"初始化翻译API失败: {e}")
            return
        
        # 解析列参数
        zh_to_en_indices = []
        en_to_zh_indices = []
        
        if args.zh2en:
            columns = parse_column_input(args.zh2en)
            for col in columns:
                try:
                    zh_to_en_indices.append(column_letter_to_index(col))
                except:
                    print(f"错误：无效的列名 '{col}'")
                    return
        
        if args.en2zh:
            columns = parse_column_input(args.en2zh)
            for col in columns:
                try:
                    en_to_zh_indices.append(column_letter_to_index(col))
                except:
                    print(f"错误：无效的列名 '{col}'")
                    return
        
        # 检查文件是否太大，推荐使用CSV模式
        try:
            file_size_mb = os.path.getsize(args.file) / (1024 * 1024)
            if file_size_mb > 50 and not args.use_csv:
                print(f"警告：文件大小为 {file_size_mb:.1f}MB，建议使用CSV模式处理大文件")
                use_csv = input("是否使用CSV模式处理？(y/n): ").strip().lower() == 'y'
            else:
                use_csv = args.use_csv
        except:
            use_csv = args.use_csv
        
        # 执行翻译
        if use_csv:
            translate_via_csv(
                args.file, 
                translators, 
                zh_to_en_indices, 
                en_to_zh_indices, 
                args.batch
            )
        else:
            translate_excel_file(
                args.file, 
                translators, 
                zh_to_en_indices, 
                en_to_zh_indices, 
                args.batch
            )

# 添加 DeepSeek 翻译器类
class DeepSeekTranslator:
    """DeepSeek-V3 翻译器实现"""
    
    def __init__(self, source='zh-CN', target='en', api_key=None, api_url=None):
        self.source = source
        self.target = target
        self.api_key = api_key
        self.api_url = api_url or "https://api.deepseek.com/v1/chat/completions"  # 替换为实际的 API 端点
        
    def translate(self, text):
        """使用 DeepSeek-V3 API 翻译文本"""
        if not text.strip():
            return ""
            
        source_lang = "中文" if "zh" in self.source else "英文"
        target_lang = "英文" if "en" in self.target else "中文"
        
        try:
            headers = {
                "Content-Type": "application/json",
                "Authorization": f"Bearer {self.api_key}"
            }
            
            payload = {
                "model": "deepseek-v3",  # 使用 DeepSeek-V3 模型
                "messages": [
                    {"role": "system", "content": f"你是一个专业翻译助手。请将下面的{source_lang}文本翻译成{target_lang}，只返回翻译结果，不要有任何解释或额外文字。"},
                    {"role": "user", "content": text}
                ],
                "temperature": 0.3  # 使用较低的温度提高翻译一致性
            }
            
            response = requests.post(self.api_url, headers=headers, data=json.dumps(payload))
            
            if response.status_code == 200:
                result = response.json()
                translation = result.get("choices", [{}])[0].get("message", {}).get("content", "")
                return translation
            else:
                print(f"翻译API错误 (代码: {response.status_code}): {response.text}")
                return text  # 失败时返回原文本
                
        except Exception as e:
            print(f"翻译过程中出错: {str(e)}")
            return text  # 异常时返回原文本

if __name__ == "__main__":
    main() 