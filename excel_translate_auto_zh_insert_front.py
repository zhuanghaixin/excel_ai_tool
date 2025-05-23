import openpyxl
from deep_translator import GoogleTranslator
import os
from tqdm import tqdm

# 初始化翻译器
t = GoogleTranslator(source='en', target='zh-CN')

# 获取当前脚本所在目录
base_dir = os.path.dirname(os.path.abspath(__file__))
input_filename = 'bbb.xlsx'  # 你可以改成英文内容的Excel
input_path = os.path.join(base_dir, input_filename)
# 自动生成输出文件名：原文件名去掉扩展名加_zh.xlsx
name, ext = os.path.splitext(input_filename)
output_filename = f'{name}_zh{ext}'
output_path = os.path.join(base_dir, output_filename)

# 加载Excel文件
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# 获取表头和数据
header = [cell.value for cell in ws[1]]
rows = list(ws.iter_rows(min_row=2, values_only=True))

# 翻译表头（英文转中文）
translated_header = []
for h in tqdm(header, desc='翻译表头'):
    if h is not None and isinstance(h, str):
        try:
            h_zh = t.translate(h)
        except Exception as e:
            h_zh = ''
    else:
        h_zh = ''
    translated_header.append(h_zh)

# 新表头：先插入翻译表头，再插入原表头
new_header = []
for h_zh, h in zip(translated_header, header):
    new_header.append(h_zh)
    new_header.append(h)

# 新数据：每列前面插入翻译
new_rows = []
for row in tqdm(rows, desc='翻译内容'):
    new_row = []
    for cell in row:
        # 翻译（空值不翻译）
        if cell is not None and isinstance(cell, str):
            try:
                translated = t.translate(cell)
            except Exception as e:
                translated = ''
        else:
            translated = ''
        new_row.append(translated)
        new_row.append(cell)
    new_rows.append(new_row)

# 写入新Excel
new_wb = openpyxl.Workbook()
new_ws = new_wb.active
new_ws.append(new_header)
for row in new_rows:
    new_ws.append(row)
new_wb.save(output_path)
print(f'翻译完成，已生成 {output_path}') 