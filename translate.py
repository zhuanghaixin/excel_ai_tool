import openpyxl
from deep_translator import GoogleTranslator, MyMemoryTranslator, BaiduTranslator
try:
    from deep_translator import deepl as DeepLTranslator
except ImportError:
    DeepLTranslator = None  # 如果导入失败就设为None
import os
from tqdm import tqdm
import string

def column_letter_to_index(column_letter):
    """将列字母转换为索引（A=0, B=1, ...）"""
    return string.ascii_uppercase.index(column_letter.upper())

# 获取当前脚本所在目录
base_dir = os.path.dirname(os.path.abspath(__file__))

# 用户输入文件名
input_filename = input("请输入要翻译的Excel文件名（如test.xlsx）: ")
if not input_filename.endswith('.xlsx'):
    input_filename += '.xlsx'
input_path = os.path.join(base_dir, input_filename)

# 检查文件是否存在
if not os.path.exists(input_path):
    print(f"错误：找不到文件 {input_filename}。请确认文件名并重试。")
    exit()

# 选择翻译API
print("\n请选择翻译API（某些地区无法使用Google翻译）:")
print("1. MyMemory翻译 (免费，无需密钥，推荐首选)")
print("2. Google翻译 (部分地区可能无法访问)")
print("3. 百度翻译 (需要API ID和密钥)")
if DeepLTranslator:
    print("4. DeepL翻译 (需要API密钥，翻译质量高)")

api_choice = input("请选择翻译API (1-4): ").strip()

# 初始化翻译器
translators = {}

if api_choice == '1':  # MyMemory翻译
    # MyMemory对于中文使用"zh-CN"，对于英文使用"en-GB"
    translators['zh_to_en'] = MyMemoryTranslator(source='zh-CN', target='en-GB')
    translators['en_to_zh'] = MyMemoryTranslator(source='en-GB', target='zh-CN')
    
elif api_choice == '2':  # Google翻译
    translators['zh_to_en'] = GoogleTranslator(source='zh-CN', target='en')
    translators['en_to_zh'] = GoogleTranslator(source='en', target='zh-CN')
    
elif api_choice == '3':  # 百度翻译
    baidu_appid = input("请输入百度翻译API的APP ID: ")
    baidu_key = input("请输入百度翻译API的密钥: ")
    
    try:
        translators['zh_to_en'] = BaiduTranslator(
            appid=baidu_appid,
            appkey=baidu_key,
            source='zh', 
            target='en'
        )
        translators['en_to_zh'] = BaiduTranslator(
            appid=baidu_appid,
            appkey=baidu_key,
            source='en',
            target='zh'
        )
    except Exception as e:
        print(f"初始化百度翻译API失败: {e}")
        exit()

elif api_choice == '4' and DeepLTranslator:  # DeepL翻译
    deepl_key = input("请输入DeepL API密钥: ")
    
    try:
        translators['zh_to_en'] = DeepLTranslator(
            api_key=deepl_key,
            source='ZH',
            target='EN'
        )
        translators['en_to_zh'] = DeepLTranslator(
            api_key=deepl_key, 
            source='EN',
            target='ZH'
        )
    except Exception as e:
        print(f"初始化DeepL API失败: {e}")
        exit()

else:
    print("无效选择，默认使用MyMemory翻译")
    translators['zh_to_en'] = MyMemoryTranslator(source='zh-CN', target='en-GB')
    translators['en_to_zh'] = MyMemoryTranslator(source='en-GB', target='zh-CN')

# 用户指定要翻译的列
print("\n请指定要翻译的列：(如不需要某方向翻译，请直接按回车跳过)")
zh_to_en_column = input("哪一列需要从【中文翻译成英文】（输入列名，如A、B等，不需要则直接按回车）: ").strip().upper()
en_to_zh_column = input("哪一列需要从【英文翻译成中文】（输入列名，如A、B等，不需要则直接按回车）: ").strip().upper()

# 验证至少有一个翻译方向
if not zh_to_en_column and not en_to_zh_column:
    print("错误：请至少指定一个翻译方向")
    exit()

# 验证列名有效性
valid_columns = list(string.ascii_uppercase)
zh_to_en_idx = None
en_to_zh_idx = None

if zh_to_en_column:
    if zh_to_en_column not in valid_columns:
        print(f"错误：'{zh_to_en_column}'不是有效的列名（A-Z）")
        exit()
    zh_to_en_idx = column_letter_to_index(zh_to_en_column)

if en_to_zh_column:
    if en_to_zh_column not in valid_columns:
        print(f"错误：'{en_to_zh_column}'不是有效的列名（A-Z）")
        exit()
    en_to_zh_idx = column_letter_to_index(en_to_zh_column)

# 输出确认信息
print(f"\n将翻译文件 {input_filename} 的：")
if zh_to_en_column:
    print(f"- {zh_to_en_column}列（从中文翻译成英文）")
if en_to_zh_column:
    print(f"- {en_to_zh_column}列（从英文翻译成中文）")
confirm = input("请确认(y/n): ").strip().lower()
if confirm != 'y':
    print("已取消操作")
    exit()

# 自动生成输出文件名
name, ext = os.path.splitext(input_filename)
output_filename = f'{name}_translated{ext}'
output_path = os.path.join(base_dir, output_filename)

# 加载Excel文件
print(f"\n正在加载 {input_filename}...")
wb = openpyxl.load_workbook(input_path)
ws = wb.active

# 获取数据的最大行列
max_row = ws.max_row
max_col = ws.max_column

# 创建新工作簿
new_wb = openpyxl.Workbook()
new_ws = new_wb.active

# 打印开始翻译的信息
translate_info = []
if zh_to_en_column:
    translate_info.append(f"{zh_to_en_column}列(中→英)")
if en_to_zh_column:
    translate_info.append(f"{en_to_zh_column}列(英→中)")
print(f"\n开始翻译{' 和 '.join(translate_info)}...")

# 处理每一行
for row_idx in tqdm(range(1, max_row + 1), desc='翻译进度'):
    new_row = []
    
    # 处理每一列
    for col_idx in range(1, max_col + 1):
        col_0_based = col_idx - 1  # 转为0-based索引
        cell_value = ws.cell(row=row_idx, column=col_idx).value
        new_row.append(cell_value)
        
        # 处理中文→英文翻译
        if zh_to_en_idx is not None and col_0_based == zh_to_en_idx:
            if cell_value is not None and isinstance(cell_value, str):
                try:
                    translated = translators['zh_to_en'].translate(cell_value)
                    if row_idx == 1:  # 如果是表头
                        translated = f"{cell_value}_en"
                except Exception as e:
                    translated = ''
                    print(f"\n翻译失败: {cell_value}, 错误: {str(e)}")
            else:
                translated = ''
            new_row.append(translated)
            
        # 处理英文→中文翻译
        elif en_to_zh_idx is not None and col_0_based == en_to_zh_idx:
            if cell_value is not None and isinstance(cell_value, str):
                try:
                    translated = translators['en_to_zh'].translate(cell_value)
                    if row_idx == 1:  # 如果是表头
                        translated = f"{cell_value}_zh"
                except Exception as e:
                    translated = ''
                    print(f"\n翻译失败: {cell_value}, 错误: {str(e)}")
            else:
                translated = ''
            new_row.append(translated)
    
    # 将行写入新表格
    new_ws.append(new_row)

# 保存新文件
new_wb.save(output_path)
print(f'\n翻译完成，已生成 {output_path}')

# 输出结果总结
if zh_to_en_column:
    print(f'- {zh_to_en_column}列：中文→英文')
if en_to_zh_column:
    print(f'- {en_to_zh_column}列：英文→中文') 